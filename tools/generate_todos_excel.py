from openpyxl import Workbook
from datetime import datetime

TODOS = [
    {
        "id": 1,
        "title": "Convertir PUML a Draw.io nativo",
        "description": "Leer HU990929-EdicionBeneficiarios.puml y generar un archivo draw.io nativo con participantes, mensajes y etiquetas. Crear HU990929-EdicionBeneficiarios.from_puml.drawio y validarlo.",
        "status": "completed",
        "estimate_hours": 3.5,
        "assumptions": "PlantUML simple; manual layout; validation passes"
    },
    {
        "id": 2,
        "title": "Validar archivo draw.io generado",
        "description": "Ejecutar el validador para comprobar ids duplicados, referencias parent/source/target rotas y nested mxCell. Corregir si es necesario.",
        "status": "completed",
        "estimate_hours": 0.5,
        "assumptions": "Validator script available; no major fixes"
    },
    {
        "id": 3,
        "title": "Decidir reemplazo en plantilla",
        "description": "Preguntar al usuario si desea que reemplace la página en `Diagrama_Secuencia 1.drawio` por la versión nativa o dejar `Diagrama_Secuencia 1.with_HU.drawio` como copia separada.",
        "status": "not-started",
        "estimate_hours": 0.25,
        "assumptions": "Decision provided quickly by user"
    },
    {
        "id": 4,
        "title": "Limpiar scripts temporales",
        "description": "Eliminar o consolidar `tools/check_drawio_temp.py` si el usuario lo aprueba; dejar `tools/check_drawio.py` como la fuente canonical del validador.",
        "status": "not-started",
        "estimate_hours": 0.5,
        "assumptions": "Small file removal or consolidation"
    },
    {
        "id": 5,
        "title": "Exportar tareas a Excel",
        "description": "Generar `tools/todos_estimacion.xlsx` con las tareas actuales y una estimación en horas (asunciones documentadas).",
        "status": "in-progress",
        "estimate_hours": 0.25,
        "assumptions": "Automation script runs locally and user approves estimates"
    }
]

HU_ACTIVITIES = [
    {
        "id": "HU-1",
        "activity": "Analizar requisitos HU990929",
        "description": "Leer la HU y entender los flujos requeridos para los diagramas de secuencia.",
        "status": "completed",
        "estimate_hours": 0.5,
        "notes": "Revisión del enunciado y aclaración de actores"
    },
    {
        "id": "HU-2",
        "activity": "Escribir PlantUML (HU990929)",
        "description": "Crear el archivo `HU990929-EdicionBeneficiarios.puml` con los tres escenarios (agregar/quitar/crear).",
        "status": "completed",
        "estimate_hours": 1.5,
        "notes": "Incluye comentarios y variantes de los flujos"
    },
    {
        "id": "HU-3",
        "activity": "Renderizar PlantUML a PNG",
        "description": "Generar PNG como fallback cuando la exportación a PDF falló por configuraciones de Java/Batik.",
        "status": "completed",
        "estimate_hours": 0.25,
        "notes": "Usado para crear versión embebida en draw.io"
    },
    {
        "id": "HU-4",
        "activity": "Crear draw.io con PNG embebido y etiquetas",
        "description": "Generar `HU990929-EdicionBeneficiarios.drawio.xml` que incluye la imagen y etiquetas separadas para evitar errores de importación.",
        "status": "completed",
        "estimate_hours": 1.0,
        "notes": "Evitar nested mxCell labels moviendo texto fuera de la imagen"
    },
    {
        "id": "HU-5",
        "activity": "Corregir import errors (mxCell)",
        "description": "Reemplazar etiquetas anidadas por mxCells independientes para que draw.io importe sin problemas.",
        "status": "completed",
        "estimate_hours": 0.5,
        "notes": "Ajustes en `HU990929-EdicionBeneficiarios.drawio.xml`"
    },
    {
        "id": "HU-6",
        "activity": "Crear versión draw.io nativa",
        "description": "Generar `HU990929-EdicionBeneficiarios.native.drawio.xml` con objetos nativos (participants, edges, labels).",
        "status": "completed",
        "estimate_hours": 2.0,
        "notes": "Page fully editable in draw.io"
    },
    {
        "id": "HU-7",
        "activity": "Insertar página nativa en plantilla copia",
        "description": "Agregar la página nativa al archivo `Diagrama_Secuencia 1.with_HU.drawio` (copia para no modificar el original).",
        "status": "completed",
        "estimate_hours": 0.5,
        "notes": "Non-destructive integration"
    },
    {
        "id": "HU-8",
        "activity": "Validar combinado y corregir página truncada",
        "description": "Ejecutar el validador, detectar página incompleta y reemplazarla por la versión completa; volver a validar.",
        "status": "completed",
        "estimate_hours": 0.75,
        "notes": "No duplicate ids after fix"
    },
    {
        "id": "HU-9",
        "activity": "Generar draw.io nativo desde PUML",
        "description": "Crear `HU990929-EdicionBeneficiarios.from_puml.drawio` con layout y edges según el PUML.",
        "status": "completed",
        "estimate_hours": 3.5,
        "notes": "Conversión automática/manual - in progress -> completed"
    },
    {
        "id": "HU-10",
        "activity": "Exportar actividades HU a Excel",
        "description": "Incluir estas actividades en el Excel de estimación solicitado por el usuario.",
        "status": "completed",
        "estimate_hours": 0.25,
        "notes": "Generado por `generate_todos_excel.py`"
    }
]


def generate_xlsx(path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Todos"

    headers = ["id", "title", "description", "status", "estimate_hours", "assumptions"]
    ws.append(headers)

    for t in TODOS:
        ws.append([t["id"], t["title"], t["description"], t["status"], t["estimate_hours"], t["assumptions"]])

    # Add generation metadata sheet
    meta = wb.create_sheet("metadata")
    meta.append(["generated_at", datetime.utcnow().isoformat() + "Z"])
    meta.append(["generator", "generate_todos_excel.py"])

    # Save to a temporary path first to avoid partial writes if target is locked
    tmp = path + ".tmp"
    wb.save(tmp)
    try:
        import os
        os.replace(tmp, path)
    except Exception:
        # Fallback: try to remove target and move
        try:
            if os.path.exists(path):
                os.remove(path)
            os.replace(tmp, path)
        except Exception as e:
            print("Failed to write final file:", e)
            raise

        # --- Create 7-day plan for HU activities ---
        capacity_per_day = 6.0  # hours per day
        days = 7
        # copy remaining estimates
        remaining = []
        for h in HU_ACTIVITIES:
            remaining.append({
                "id": h["id"],
                "activity": h["activity"],
                "remaining": float(h.get("estimate_hours", 0.0)),
            })

        schedule = {d: [] for d in range(1, days + 1)}
        for day in range(1, days + 1):
            available = capacity_per_day
            i = 0
            while i < len(remaining) and available > 0.0:
                item = remaining[i]
                if item["remaining"] <= 0:
                    i += 1
                    continue
                take = min(item["remaining"], available)
                schedule[day].append({
                    "id": item["id"],
                    "activity": item["activity"],
                    "hours": take,
                })
                item["remaining"] -= take
                available -= take
                if item["remaining"] <= 0:
                    i += 1

        plan_path = path.replace('.xlsx', '_plan.xlsx')
        wb2 = Workbook()
        ws_p = wb2.active
        ws_p.title = "Plan 7 dias"
        ws_p.append(["day", "id", "activity", "hours"])
        for d in range(1, days + 1):
            entries = schedule[d]
            if not entries:
                ws_p.append([d, "", "(no tasks)", 0])
            else:
                for e in entries:
                    ws_p.append([d, e["id"], e["activity"], e["hours"]])

        ws_s = wb2.create_sheet("Resumen")
        total_hours = sum(h["estimate_hours"] for h in HU_ACTIVITIES)
        ws_s.append(["total_hu_hours", total_hours])
        ws_s.append(["capacity_per_day", capacity_per_day])
        ws_s.append(["days", days])
        ws_s.append(["estimated_days_needed", round(total_hours / capacity_per_day, 2)])

        wb2.save(plan_path)
        print("Wrote HU plan to:", plan_path)

    # Write HU activities into a separate sheet
    ws2 = wb.create_sheet("HU_Actividades")
    headers2 = ["id", "activity", "description", "status", "estimate_hours", "notes"]
    ws2.append(headers2)
    for h in HU_ACTIVITIES:
        ws2.append([h["id"], h["activity"], h["description"], h["status"], h["estimate_hours"], h["notes"]])

    # Save again after adding HU sheet
    wb.save(tmp)
    try:
        import os
        os.replace(tmp, path)
    except Exception:
        try:
            if os.path.exists(path):
                os.remove(path)
            os.replace(tmp, path)
        except Exception as e:
            print("Failed to write final file:", e)
            raise


if __name__ == "__main__":
    out = "d:/Repositorios/VASS-BHD/Diagramas-HU-BHDIB/Diagramas-HU-BHDIB/tools/todos_estimacion.xlsx"
    generate_xlsx(out)
    print("Wrote:", out)
